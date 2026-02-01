import re
import time
import json
from datetime import datetime
import serial
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

"""
IMPORTANT: You cannot have the serial monitor open when running this!
Close it before running or else its not gon work!
To Run:
1. Run Arduino Code
2. In a terminal, go into directory with this file and type "python3 serial_to_excel.py"
3. To stop logging, Ctrl + C
"""

PORT = "/dev/ttyACM0"      # for windows it would look like "COM3"
BAUD = 115200 # stable gps baud
OUTPUT_XLSX = "serial_log_2.xlsx" # change this to name the output file to the antenna name
SHEET_NAME = "log"
FLUSH_EVERY_N_ROWS = 25    # Save file every N lines
READ_TIMEOUT_SEC = 1


KV_PAIR_RE = re.compile(r"^\s*([A-Za-z_][A-Za-z0-9_]*)\s*=\s*(.+?)\s*$")

def parse_line(line: str):
    """
    Returns a dict of parsed fields if possible, else None.
    Supports:
      - key=value, key2=value2 (comma separated)
      - CSV values (comma separated)
    """
    line = line.strip()
    if not line:
        return None

    # try "key=value" style 
    parts = [p.strip() for p in line.split(",")]
    kv = {}
    all_kv = True
    for p in parts:
        m = KV_PAIR_RE.match(p)
        if not m:
            all_kv = False
            break
        kv[m.group(1)] = m.group(2)

    if all_kv and kv:
        return kv

    # Try CSV style if it contains commas
    if "," in line:
        return {f"col_{i+1}": v.strip() for i, v in enumerate(parts)}

    return None

def parse_line(line: str):
    """
    Returns a dict of parsed fields if possible, else None.
    Supports:
      - JSON objects
      - key=value pairs
      - CSV values
    """
    line = line.strip()
    if not line:
        return None

    # json
    if line.startswith("{") and line.endswith("}"):
        try:
            data = json.loads(line)
            return data if isinstance(data, dict) else None
        except json.JSONDecodeError:
            return None

    # key=value
    parts = [p.strip() for p in line.split(",")]
    kv = {}
    for p in parts:
        if "=" not in p:
            kv = None
            break
        k, v = p.split("=", 1)
        kv[k.strip()] = v.strip()

    if kv:
        return kv

    # csv
    if "," in line:
        return {f"col_{i+1}": v.strip() for i, v in enumerate(parts)}

    return None

def autosize_columns(ws, max_width=60):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        best = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                continue
            best = max(best, len(str(val)))
        ws.column_dimensions[letter].width = min(max(10, best + 2), max_width)

def main():
    ser = serial.Serial(PORT, BAUD, timeout=READ_TIMEOUT_SEC)
    time.sleep(2)  # give Arduino time to reset on open

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    headers = ["timestamp", "raw"]
    header_set = set(headers)

    ws.append(headers)

    rows_since_save = 0

    print(f"Logging serial from {PORT} @ {BAUD} to {OUTPUT_XLSX}")
    print("Press Ctrl+C to stop.\n")

    try:
        while True:
            raw = ser.readline().decode(errors="replace").strip()
            if not raw:
                continue

            ts = datetime.now().isoformat(timespec="seconds")

            parsed = parse_line(raw)

            if parsed:
                new_keys = [k for k in parsed.keys() if k not in header_set]
                if new_keys:
                    for k in new_keys:
                        headers.append(k)
                        header_set.add(k)
                    for i, h in enumerate(headers, start=1):
                        ws.cell(row=1, column=i, value=h)

            row = {h: "" for h in headers}
            row["timestamp"] = ts
            row["raw"] = raw
            if parsed:
                for k, v in parsed.items():
                    row[k] = v

            ws.append([row[h] for h in headers])
            rows_since_save += 1

            print(f"{ts}  {raw}")

            if rows_since_save >= FLUSH_EVERY_N_ROWS:
                autosize_columns(ws)
                wb.save(OUTPUT_XLSX)
                rows_since_save = 0

    except KeyboardInterrupt:
        print("\nStopping...")

    finally:
        autosize_columns(ws)
        wb.save(OUTPUT_XLSX)
        ser.close()
        print(f"Saved: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
